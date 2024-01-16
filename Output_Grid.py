
# coding: utf-8
from __future__ import division
# Importations
import os 
import numpy as np
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import *
from openpyxl.styles import Font, Color
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

'''
NOT CURRENTLY NEEDED!!!! DO NOT NEED TO TRANSLATE
Returns          void

Inputs:
TotalYears       double
Cells_Obj        array of objects from class Cells
NCells           int
BurnCells_Set    int set
Folder           string
Sim              int
spotting         boolean
verbose          boolearn
'''
# Generates scenario.dat files for optimization (to be modified, not efficient)
def ScenarioOutput(TotalYears, Cells_Obj, NCells, BurntCells_Set, Folder, Sim, spotting, verbose):
    # FI and FS Debug
    # Parameters and sets for the stochastic model
    HPeriods = set()
    FPeriods = set()
    HPeriodsList = []
    FPeriodsList = [[] for i in range(0,TotalYears)]
    lenFP = 0
    lenHP = 0

    MaxTF = [0 for i in range(0,TotalYears)]
    MinTF = [99999 for i in range(0,TotalYears)]

    for c in Cells_Obj.keys():
        for t in range(0,2016):
            if Cells_Obj[c].FICell[t] == 1:
                MinTF[Cells_Obj[c].FireStartsSeason-1] = t+1
                if (t+1) > MaxTF[Cells_Obj[c].FireStartsSeason-1]:
                    MaxTF[Cells_Obj[c].FireStartsSeason-1] = t+1
                        
            if len(Cells_Obj[c].FSCell[t]) >= 1:
                if (t+1) > MaxTF[Cells_Obj[c].FireStartsSeason-1]:
                    MaxTF[Cells_Obj[c].FireStartsSeason-1] = t+1
                                
    for y in range(0,len(MaxTF)):
        if MaxTF[y] < MinTF[y]:
            MinTF[y] = 0
        
    Special = 0
    for y in range(0,TotalYears):
        if y == 0:
            HPeriods = HPeriods.union(set([1]))
            for p in range(2, MaxTF[y] - MinTF[y] + 1 + 2):
                FPeriods = FPeriods.union(set([p]))
                FPeriodsList[y] = FPeriods
                
        
        if y >= 1:
            if max(FPeriods)+1 in HPeriods:
                HPeriods = HPeriods.union(set([max(HPeriods)+1]))
            else:
                HPeriods = HPeriods.union(set([max(FPeriods)+1]))
            
            if MaxTF[y] != MinTF[y] and MaxTF[y] >= MinTF[y]:
                lenFP = max(HPeriods)
    
                for p in range(lenFP + 1, lenFP + 1 + MaxTF[y] - MinTF[y] + 1):
                    if p not in HPeriods:
                        FPeriods = FPeriods.union(set([p]))
                            
            if MaxTF[y] == MinTF[y]:
                if p not in HPeriods:
                    FPeriods = FPeriods.union(set([max(HPeriods)+1]))
                    
                                    
            FPeriodsList[y] = FPeriods
        
    #Printing forest's data to a txt file
    if spotting == True:
        Folder = os.path.join(Folder, "ScenariosSpotting")
    else:
        Folder = os.path.join(Folder, "Scenarios")
    if not os.path.exists(Folder):
        os.makedirs(Folder)
    filed = open(os.path.join(Folder, "Scenario"+str(Sim)+".dat"),"w")
        
    HPeriodaux = [i for i in HPeriods]
    HPeriodaux.sort()
            
    if verbose == True:
        print("HPeriods:", HPeriodaux)
        print("FPeriods:", sorted(FPeriods))
        
    filed.write("#########################################################################################################################################################################################################"+"\n")
    filed.write("#																																																		   																																																															 "+"\n")
    filed.write("#                                                      Sets																																																																																																		 "+"\n")
    filed.write("#																																																																																																																	 "+"\n")							
    filed.write("#########################################################################################################################################################################################################"+"\n")
        
    filed.write("set HPeriods:= ")
    for i in sorted(HPeriods):
        if max(HPeriods) != i:
            filed.write(str(i)+ " ")
        else:
            filed.write(str(i))
    filed.write(";"+"\n"+"\n")
        
    filed.write("set FPeriods:= ")
    for i in sorted(FPeriods):
        if max(FPeriods) != i:
            filed.write(str(i)+ " ")
        else:
            filed.write(str(i))
            
    filed.write(";"+"\n"+"\n")
                
    filed.write("#########################################################################################################################################################################################################"+"\n")
    filed.write("#																																																		   																																																															 "+"\n")
    filed.write("#                                                      Parameters																																																																																																		 "+"\n")
    filed.write("#																																																																																																																	 "+"\n")							
    filed.write("#########################################################################################################################################################################################################"+"\n")
                    
    # DLW Feb 2018: this is writing a Param for Pyomo or AMPL
    # call FI to a file. It is not the same as the statistic FI
        # gethered in the simulator, but it is related.
    if verbose == True:
        print("\n", "(AMPL/Pyomo) param FI:=")
    filed.write("param FI:="+"\n")
    for c in BurntCells_Set:
        for y in range(0,TotalYears):
            if MinTF[y] <2016:
                if Cells_Obj[c-1].FICell[MinTF[y]-1] == 1 and (Cells_Obj[c-1].FireStartsSeason-1) == y:
                        if verbose == True:
                            print(c, HPeriodaux[y]+1, 1 )
                        filed.write(str(c)+" "+str(HPeriodaux[y]+1)+" "+str(1)+"\n")
                        break;
    if verbose == True:
        print(";")
    filed.write(";"+"\n"+"\n")
    filed.write("")
        
    #Printing FS
    for y in range(0,TotalYears):
        if TotalYears - y >= 2:
            FPeriodsList[TotalYears-1-y] = FPeriodsList[TotalYears-1-y].difference(FPeriodsList[TotalYears-1-y-1]) 

        # The ampl/pyomo param FS is related to the simulator statistic FS.
    if verbose == True:
        print("\n","(AMPL/Pyomo) param FS:=")
    filed.write("param FS:="+"\n")
    for y in range(0,TotalYears):
        for c in Cells_Obj.keys():
            for t in range(0,2016):
                if (Cells_Obj[c].FireStartsSeason-1) == y and len(Cells_Obj[c].FSCell[t]) >= 1:
                    for j in range(0,len(Cells_Obj[c].FSCell[t])):
                        if len(FPeriodsList[y]) > 0:
                            P = t - MinTF[y] + min(FPeriodsList[y]) +1
                            if verbose == True:
                                print(Cells_Obj[c].FSCell[t][j],c+1,P,1)
                            filed.write(str(Cells_Obj[c].FSCell[t][j])+" "+str(c+1)+" "+str(P)+" "+str(1)+"\n")
    if verbose == True:
        print(";","\n")
    filed.write(";"+"\n")
                                    
    filed.close()

    

'''
Returns          void

Inputs:
Cells_Obj        dictionary of objects from class Cells
Sim              int
Folder           string
heuristic        boolean
TotalSims        int
TotalYears       int
NCells           int 
BurntCells_Set   int set
AvailCells_Set   int set
spotting         boolean
'''
# Generates excel file with statistics 
def ExcelOutput(Cells_Obj, Sim, Folder, heuristic, TotalSims, TotalYears, NCells, 
                BurntCells_Set, AvailCells_Set, HarvestCells_Set, spotting):

    ## Step 1: Statistics
    # Copy of simulation value
    Sim2=Sim
    Sim = 1
    
    # Initialize arrays of arrays for statistics
    GlobalStats = [[] for i in range(0, Sim)]
    CellsNHFStats = [[] for i in range(0, NCells)]
    AveragesAvailNH = [[] for i in range(0, Sim)]
    AveragesFireNH = [[] for i in range(0, Sim)]

    # If harvesting heuristics are applied, initialize empty arrays of arrays
    if heuristic != 0:
        CellsHHStats = [[] for i in range(0,NCells)]
        CellsHFStats = [[] for i in range(0,NCells)]
        AveragesAvailH = [[] for i in range(0,TotalSims)]
        AveragesFireH = [[] for i in range(0,TotalSims)]
        AveragesHarvestedH = [[] for i in range(0,TotalSims)]

    # Percentages
    AveragesAvailNH[Sim - 1] = np.round(len(AvailCells_Set) / NCells * 100.0, 2)
    AveragesFireNH[Sim - 1] = np.round(len(BurntCells_Set) / NCells * 100.0, 2)
    #AveragesHarvest[Sim - 1] = np.round(len(HarvestCells_Set) / NCells * 100.0, 2)
    if heuristic != 0:
        AveragesHarvestedH[Sim - 1] = np.round(len(HarvestCells_Set) / NCells * 100.0, 2)
        AveragesAvailH[Sim - 1] = np.round(len(AvailCells_Set) / NCells * 100.0, 2)
        AveragesFireH[Sim - 1] = np.round(len(BurntCells_Set) / NCells * 100.0, 2)
        
    # Burnt no heuristic
    for br in BurntCells_Set:
        CellsNHFStats[br-1].append(Cells_Obj[br-1].FireStartsSeason)     
    
    # If heuristic
    if heuristic != 0:
        SimResults = [CellsNHFStats,CellsHFStats,CellsHHStats]
    else:
        SimResults = [CellsNHFStats]
            
    GlobalStats[Sim - 1] = SimResults
        
    ## Step 2: Excel
    # To excel
    wb1 = Workbook()
    auxrow = 2
    ws1 = wb1.active
    
    ws1.title = "Global Results"
    ws1.cell(row=1, column=1, value = "Heuristic") 
    #ws1.cell(row=1, column=2, value = "No Heuristic")
    
    if heuristic == True:
        ws1.cell(row = auxrow, column = 1, value = len(BurntCells_Set))
    #ws1.cell(row = auxrow, column = 2, value = len(BurntCells_Set))
    auxrow += 1
            
    font = Font(name='Calibri',
                 size=11,
                 bold=True,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000')
    
    # If heuristic, add harvesting statistics
    if heuristic != 0:
        ws1.cell(row=1, column=4, value = "AVG Heuristic") 
        ws1.cell(row=4, column=4, value = "STD Heuristic") 
        ws1.cell(row=7, column=5, value = "Heuristic") 
        
        FormulaAVGH = "=AVERAGE(A2:A" + str(TotalSims + 1) + ")"
        FormulaSTDH = "=STDEV(A2:A" + str(TotalSims + 1) + ")"
        FormulaT = "=(E2-D2)/SQRT((D5^2+E5^2)/" + str(TotalSims) + ")"
        FormulaGL = "=ROUND((" + str(TotalSims - 1) + ")*((D5^2+E5^2)^2)/(D5^4+E5^4),0)"
        FormulaTest = "=TDIST(G2,H2,2)"
        FormulaResult = "=IF(I2<0.05,\"Results are different 95% \", \"Avgs are identical\")"
        
        ws1.cell(row=1, column=7, value = "TestT") 
        ws1.cell(row=1, column=8, value = "G.L")
        ws1.cell(row=1, column=9, value = "PValue")
        ws1.cell(row=1, column=10, value = "Result (95%)")

        ws1.cell(row=2, column=4, value = FormulaAVGH) 
        ws1.cell(row=5, column=4, value = FormulaSTDH) 
        ws1.cell(row=2, column=7, value = FormulaT)
        ws1.cell(row=2, column=8, value = FormulaGL)
        ws1.cell(row=2, column=9, value = FormulaTest)
        ws1.cell(row=2, column=10, value = FormulaResult)

    ws1.cell(row=1, column=5, value = "AVG No Heuristic")
    ws1.cell(row=4, column=5, value = "STD No Heuristic")
    ws1.cell(row=7, column=6, value = "No Heuristic")
    ws1.cell(row=8, column=4, value = "% Available")
    ws1.cell(row=9, column=4, value = "% Harvested")
    ws1.cell(row=10, column=4, value = "% Burnt")
        
    FormulaAVGNH = "=AVERAGE(B2:B" + str(TotalSims + 1) + ")"
    FormulaSTDNH = "=STDEV(B2:B" + str(TotalSims + 1) + ")"
        
    ws1.cell(row=2, column=5, value = FormulaAVGNH)
    ws1.cell(row=5, column=5, value = FormulaSTDNH)
    
    bold = ws1.row_dimensions[1]
    bold.font = Font(bold=True)
        
    # Second sheet: Simulation detailed results
    ws2 = wb1.create_sheet() # insert at the end (default)

    ws2.title = "Harvest-Fire Details"
    ws2.cell(row=1, column=1, value = "Heuristic") 
    ws2.cell(row=2, column=1, value = "Cell/HPeriod") 
    ws2.cell(row=1, column=13, value = "No Heuristic")
    ws2.cell(row=2, column=7, value = "Cell/FPeriod")
    ws2.cell(row=2, column=13, value = "Cell/FPeriod")
    
    #Seasons and cells
    for t in range(1,TotalYears + 1):
        ws2.cell(row=2, column=t+1, value = t)
        ws2.cell(row=2, column=t+7, value = t)
        ws2.cell(row=2, column=t+13, value = t)
        
    for c in range(1,NCells + 1):
        ws2.cell(row=c+2, column=1, value = c)
        ws2.cell(row=c+2, column=7, value = c)
        ws2.cell(row=c+2, column=13, value = c)
        
    # Third sheet: AVG details
    ws3 = wb1.create_sheet() # insert at the end (default)	
    ws3.title = "AVG Statistics"
    
    ws3.cell(row=1, column=1, value = "Heuristic") 
    ws3.cell(row=2, column=2, value = "AVG HPeriod")
    ws3.cell(row=1, column=11, value = "No Heuristic")
    ws3.cell(row=2, column=1, value = "Cell") 
    ws3.cell(row=2, column=3, value = "Frequency")
    ws3.cell(row=2, column=4, value = "Probability")
    ws3.cell(row=2, column=6, value = "Cell") 
    ws3.cell(row=2, column=7, value = "AVG FPeriod")
    ws3.cell(row=2, column=8, value = "Frequency")
    ws3.cell(row=2, column=9, value = "Probability")
    ws3.cell(row=2, column=11, value = "Cell") 
    ws3.cell(row=2, column=12, value = "AVG FPeriod")
    ws3.cell(row=2, column=13, value = "Frequency")
    ws3.cell(row=2, column=14, value = "Probability")
        
    # Cells
    for c in range(1,NCells + 1):
        ws3.cell(row=c+2, column=1, value = c)
        ws3.cell(row=c+2, column=6, value = c)
        ws3.cell(row=c+2, column=11, value = c)
        
    # Compute statistics
    if heuristic != 0:
        TotalHFires = np.zeros(NCells)   #[0 for i in range(0,NCells)]
        TotalHHarvest = np.zeros(NCells) #[0 for i in range(0,NCells)]
        AVGHFPeriod = np.zeros(NCells)   # [0 for i in range(0,NCells)]
        AVGHHPeriod = np.zeros(NCells)   # [0 for i in range(0,NCells)]
        ProbHFire = np.zeros(NCells)     #[0 for i in range(0,NCells)]
        ProbHHarvest = np.zeros(NCells)  #[0 for i in range(0,NCells)]
    
    TotalNHFires = np.zeros(NCells)      # [0 for i in range(0,NCells)]
    AVGNHFPeriod = np.zeros(NCells)      # [0 for i in range(0,NCells)]
    ProbNHFire = np.zeros(NCells)        # [0 for i in range(0,NCells)]
    
    for c in range(0,NCells):
        TotalNHFires[c] += len(GlobalStats[Sim-1][0][c])
        
        if heuristic != 0:
            TotalHFires[c] += len(GlobalStats[TotalSims-1][1][c])
            TotalHHarvest[c] += len(GlobalStats[TotalSims-1][2][c])
            ProbHFire[c] = TotalHFires[c] / TotalSims
            ProbHHarvest[c] = TotalHHarvest[c] / TotalSims

        ProbNHFire[c] = TotalNHFires[c] / Sim

        if TotalNHFires[c] != 0:
            AVGNHFPeriod[c] = sum(GlobalStats[Sim-1][0][c]) / TotalNHFires[c]
        else:
            AVGNHFPeriod[c] = 0
        
        if heuristic != 0:
            if TotalHFires[c] != 0:
                AVGHFPeriod[c] = sum(GlobalStats[Sim-1][1][c]) / TotalHFires[c]
            else:
                AVGHFPeriod[c] = 0
                
            if TotalHHarvest[c] != 0:
                AVGHHPeriod[c] = sum(GlobalStats[Sim-1][2][c]) / TotalHHarvest[c]
            else:
                AVGHHPeriod[c]  = 0

        ws3.cell(row=c+3, column=12, value = AVGNHFPeriod[c])
        ws3.cell(row=c+3, column=13, value = TotalNHFires[c])
        ws3.cell(row=c+3, column=14, value = ProbNHFire[c])

        if heuristic != 0:
            ws3.cell(row=c+3, column=7, value = AVGHFPeriod[c])
            ws3.cell(row=c+3, column=8, value = TotalHFires[c])
            ws3.cell(row=c+3, column=9, value = ProbHFire[c])
            ws3.cell(row=c+3, column=2, value = AVGHHPeriod[c])
            ws3.cell(row=c+3, column=3, value = TotalHHarvest[c])
            ws3.cell(row=c+3, column=4, value = ProbHHarvest[c])
            
        verbose = False
        if verbose == True:
            print("---------------------------------------- Cell",c+1, "Statistics ----------------------------------------")
            print("\n","Cell", c+1,"No Heur Fire:", GlobalStats[Sim-1][0][c])
            print("Total NH Fires",c+1,"sim",Sim-1,":", TotalNHFires[c])
            
            print("Prob of NH fire",c+1,":", ProbNHFire[c] )
            print("AVG NH Fire period: ", AVGNHFPeriod[c],"\n")
            
            if heuristic != 0:
                print("Cell", c+1,"Heur Fire:", GlobalStats[Sim-1][1][c])
                print("Total H Fires",c+1,"sim",Sim-1,":", TotalHFires[c])
                print("Prob of H fire",c+1,":", ProbHFire[c] )
                print("AVG H Fire period: ", AVGHFPeriod[c],"\n")
                print("Cell", c+1,"Heur Harvested:", GlobalStats[Sim-1][2][c])
                print("Total H Harvest",c+1,"sim",Sim-1,":", TotalHHarvest[c])
                print("Prob of H Harvest",c+1,":", ProbHHarvest[c] )
                print("AVG H Harvest period: ", AVGHHPeriod[c],"\n")
            
            print("Average of Available cells no heur: ", np.mean(AveragesAvailNH))
            
            if heuristic != 0:
                print("Average of Available cells heur: ", np.mean(AveragesAvailH))
            print("Average of Burnt cells no heur: ", np.mean(AveragesFireNH))
            
            if heuristic != 0:
                print("Average of Burnt cells heur: ", np.mean(AveragesFireH))
                print("Average of Harvested cells heur: ", np.mean(AveragesHarvestedH))

    if heuristic != 0:		
        ws1.cell(row=8, column=5, value = np.mean(AveragesAvailH))
        ws1.cell(row=9, column=5, value =  np.mean(AveragesHarvestedH))
        ws1.cell(row=10, column=5, value = np.mean(AveragesFireH))
    
    ws1.cell(row=8, column=6, value = np.mean(AveragesAvailNH))
    ws1.cell(row=9, column=6, value =  0)
    ws1.cell(row=10, column=6, value = np.mean(AveragesFireNH))
    
    #Detailed frequency
    for c in range(0, NCells):
        for t in range(0, TotalYears):
            if heuristic != 0:
                ws2.cell(row=c+3, column=2+t, value = GlobalStats[Sim-1][2][c].count(t+1))
                ws2.cell(row=c+3, column=8+t, value = GlobalStats[Sim-1][1][c].count(t+1))
                ws2.cell(row=c+3, column=14+t, value = GlobalStats[Sim-1][0][c].count(t+1))
    
    # If spotting is active
    if spotting == True:
        Folder = os.path.join(Folder, "StatisticsSpotting")
    else:
        Folder = os.path.join(Folder, "Statistics")
        
    if not os.path.exists(Folder):
        os.makedirs(Folder)
    wb1.save(os.path.join(Folder, "Scenario"+str(Sim2)+".xlsx"))
    
                      


                      

'''
Returns          void

Inputs:
Folder           string
Rows             int
Cols             int
BurnCells_Set    int set
Sim              int
spotting         boolean
verbose          boolean
cellsize         double
'''                      
# Generate the grid: 0s (non-burnt) and 1s (burnt) for checking forest status                  
def OutputGrid(Folder, Rows, Cols, BurntCells_Set, Sim, spotting, verbose, cellsize):
    if spotting == True:
        Folder = os.path.join(Folder, "ForestGridsSpotting")
    else:
        Folder = os.path.join(Folder, "ForestGrids")
        
    if not os.path.exists(Folder):
        os.makedirs(Folder)
    filed = open(os.path.join(Folder, "ForestGrid"+str(Sim)+".csv"),"w")
    filed.write("ncols "+str(Cols)+"\n")
    filed.write("nrows "+str(Rows)+"\n")
    filed.write("xllcorner 457900\n")
    filed.write("yllcorner 5716800\n")
    filed.write("cellsize "+str(cellsize)+"\n")
    filed.write("NODATA_value -9999\n")

    NcellAux = 1
    for r in range(1,Rows+1):
        for c in range(1,Cols+1):
            if NcellAux in BurntCells_Set and c < Cols:
                filed.write("1 ")
            
            if NcellAux not in BurntCells_Set and c < Cols:
                filed.write("0 ")
                      
            if NcellAux in BurntCells_Set and c==Cols:
                filed.write("1\n")
            
            if NcellAux not in BurntCells_Set and c == Cols:
                filed.write("0\n")
                        
            NcellAux += 1
    

